import io
import re
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field, field_validator
from typing import Optional
from utils.auth import require_admin
from db import supabase
from routers.attendance import effective_work_minutes

_TIME_RE = re.compile(r'^([01]\d|2[0-3]):[0-5]\d$')


class CompanyUpdateRequest(BaseModel):
    name: Optional[str] = Field(default=None, max_length=100)
    address: Optional[str] = Field(default=None, max_length=255)
    lat: Optional[float] = Field(default=None, ge=-90, le=90)
    lng: Optional[float] = Field(default=None, ge=-180, le=180)
    radius_meters: Optional[int] = Field(default=None, ge=10, le=50000)
    work_start: Optional[str] = None  # "HH:MM"
    work_end: Optional[str] = None    # "HH:MM"
    work_saturday: Optional[bool] = None
    work_sunday: Optional[bool] = None
    flexible_attendance: Optional[bool] = None
    min_work_minutes: Optional[int] = Field(default=None, ge=30, le=1440)
    multi_location: Optional[bool] = None
    overtime_enabled: Optional[bool] = None

    @field_validator("work_start", "work_end")
    @classmethod
    def validate_time(cls, v):
        if v is not None and not _TIME_RE.match(v):
            raise ValueError("Format waktu harus HH:MM (contoh: 08:00)")
        return v


class AttendanceCorrectionRequest(BaseModel):
    clock_in: Optional[str] = None   # ISO datetime string, e.g. "2026-05-05T08:00:00"
    clock_out: Optional[str] = None  # ISO datetime string
    break_start: Optional[str] = None
    break_end: Optional[str] = None
    notes: Optional[str] = None

router = APIRouter(prefix="/admin", tags=["admin"])

_STATUS_ID = {
    "present":     "Hadir",
    "late":        "Terlambat",
    "early_leave": "Pulang Cepat",
    "absent":      "Absen",
    "leave":       "Cuti",
    "pending":     "Menunggu",
    "approved":    "Disetujui",
    "rejected":    "Ditolak",
}
_LEAVE_TYPE_ID = {
    "annual":   "Tahunan",
    "sick":     "Sakit",
    "personal": "Keperluan Pribadi",
    "other":    "Lainnya",
}


def _make_xlsx(sheet_title: str, headers: list, rows: list, fill_hex: str = "2563EB") -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    hfill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF", size=10)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18

    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=val)

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max(max_len + 2, 8), 42)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _xlsx_response(buf: io.BytesIO, filename: str) -> StreamingResponse:
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/stats")
async def get_stats(admin: dict = Depends(require_admin)):
    """Dashboard stats for today."""
    from datetime import datetime
    from zoneinfo import ZoneInfo
    today = datetime.now(ZoneInfo("Asia/Jakarta")).date().isoformat()
    company_id = admin["company_id"]

    # Total employees
    emp_res = supabase.table("profiles").select("id", count="exact").eq("company_id", company_id).eq("is_active", True).execute()

    # Today attendance
    att_res = supabase.table("attendance").select("id,status", count="exact").eq("company_id", company_id).eq("date", today).execute()

    # Pending leaves
    leave_res = supabase.table("leave_requests").select("id", count="exact").eq("company_id", company_id).eq("status", "pending").execute()

    # Pending overtime
    ot_res = supabase.table("overtime_requests").select("id", count="exact").eq("company_id", company_id).eq("status", "pending").execute()

    # Pending corrections
    corr_res = supabase.table("attendance_corrections").select("id", count="exact").eq("company_id", company_id).eq("status", "pending").execute()

    return {
        "total_employees": emp_res.count or 0,
        "present_today": att_res.count or 0,
        "pending_leaves": leave_res.count or 0,
        "pending_overtime": ot_res.count or 0,
        "pending_corrections": corr_res.count or 0,
    }


@router.get("/attendance")
async def admin_attendance(
    page: int = 1,
    per_page: int = 10,
    date_from: str = None,
    date_to: str = None,
    admin: dict = Depends(require_admin),
):
    if date_from and date_to and date_from > date_to:
        raise HTTPException(400, "date_from tidak boleh setelah date_to")
    per_page = min(per_page, 100)
    offset = (page - 1) * per_page
    q = (
        supabase.table("attendance")
        .select("*, profiles(full_name,position), locations(name)", count="exact")
        .eq("company_id", admin["company_id"])
        .order("date", desc=True)
        .order("clock_in", desc=True)
    )
    if date_from:
        q = q.gte("date", date_from)
    if date_to:
        q = q.lte("date", date_to)

    res = q.range(offset, offset + per_page - 1).execute()
    return {"data": res.data, "total": res.count, "page": page, "per_page": per_page}


@router.get("/leave")
async def admin_leave(
    page: int = 1,
    per_page: int = 10,
    status_filter: str = None,
    date_from: str = None,
    date_to: str = None,
    admin: dict = Depends(require_admin),
):
    if date_from and date_to and date_from > date_to:
        raise HTTPException(400, "date_from tidak boleh setelah date_to")
    per_page = min(per_page, 100)
    offset = (page - 1) * per_page
    q = (
        supabase.table("leave_requests")
        .select("*, profiles!leave_requests_user_id_fkey(full_name,position)", count="exact")
        .eq("company_id", admin["company_id"])
        .order("created_at", desc=True)
    )
    if status_filter:
        q = q.eq("status", status_filter)
    if date_from:
        q = q.gte("start_date", date_from)
    if date_to:
        q = q.lte("start_date", date_to)

    res = q.range(offset, offset + per_page - 1).execute()
    return {"data": res.data, "total": res.count, "page": page, "per_page": per_page}


@router.get("/overtime")
async def admin_overtime(
    page: int = 1,
    per_page: int = 10,
    status_filter: str = None,
    date_from: str = None,
    date_to: str = None,
    admin: dict = Depends(require_admin),
):
    if date_from and date_to and date_from > date_to:
        raise HTTPException(400, "date_from tidak boleh setelah date_to")
    per_page = min(per_page, 100)
    offset = (page - 1) * per_page
    q = (
        supabase.table("overtime_requests")
        .select("*, profiles!overtime_requests_user_id_fkey(full_name,position)", count="exact")
        .eq("company_id", admin["company_id"])
        .order("created_at", desc=True)
    )
    if status_filter:
        q = q.eq("status", status_filter)
    if date_from:
        q = q.gte("date", date_from)
    if date_to:
        q = q.lte("date", date_to)

    res = q.range(offset, offset + per_page - 1).execute()
    return {"data": res.data, "total": res.count, "page": page, "per_page": per_page}


@router.get("/employees")
async def admin_employees(
    page: int = 1,
    per_page: int = 10,
    search: str = None,
    is_active: bool = None,
    admin: dict = Depends(require_admin),
):
    import asyncio
    from utils.auth import _get_auth_email

    per_page = min(per_page, 100)
    offset = (page - 1) * per_page
    q = (
        supabase.table("profiles")
        .select("*", count="exact")
        .eq("company_id", admin["company_id"])
        .order("full_name")
    )
    if search:
        q = q.ilike("full_name", f"%{search}%")
    if is_active is not None:
        q = q.eq("is_active", is_active)
    res = q.range(offset, offset + per_page - 1).execute()

    rows = res.data or []
    if rows:
        emails = await asyncio.gather(
            *[_get_auth_email(r["id"]) for r in rows],
            return_exceptions=True,
        )
        for row, email in zip(rows, emails):
            row["email"] = email if isinstance(email, str) else None

    return {"data": rows, "total": res.count, "page": page, "per_page": per_page}


@router.get("/employees/{user_id}/attendance")
async def employee_attendance_history(
    user_id: str,
    page: int = 1,
    per_page: int = 20,
    admin: dict = Depends(require_admin),
):
    try:
        emp = supabase.table("profiles").select("company_id, full_name").eq("id", user_id).single().execute()
    except Exception:
        raise HTTPException(404, "Karyawan tidak ditemukan")
    if not emp.data or emp.data["company_id"] != admin["company_id"]:
        raise HTTPException(404, "Karyawan tidak ditemukan")

    per_page = min(per_page, 100)
    offset = (page - 1) * per_page
    res = (
        supabase.table("attendance")
        .select("*", count="exact")
        .eq("user_id", user_id)
        .eq("company_id", admin["company_id"])
        .order("date", desc=True)
        .range(offset, offset + per_page - 1)
        .execute()
    )
    return {
        "employee": emp.data,
        "data": res.data,
        "total": res.count,
        "page": page,
        "per_page": per_page,
    }


@router.get("/company")
async def get_company(admin: dict = Depends(require_admin)):
    try:
        res = (
            supabase.table("companies")
            .select("*")
            .eq("id", admin["company_id"])
            .single()
            .execute()
        )
    except Exception:
        raise HTTPException(404, "Company not found")
    if not res.data:
        raise HTTPException(404, "Company not found")
    return res.data


@router.patch("/company")
async def update_company(
    body: CompanyUpdateRequest,
    admin: dict = Depends(require_admin),
):
    updates = {k: v for k, v in body.model_dump().items() if v is not None}
    if not updates:
        raise HTTPException(400, "No fields to update")

    company_id = admin["company_id"]

    # Detect first-time multi_location toggle ON → seed initial location from
    # existing lat/lng and assign all current employees so nothing breaks.
    try:
        current = supabase.table("companies").select(
            "multi_location, lat, lng, radius_meters, name"
        ).eq("id", company_id).single().execute()
        current_data = current.data or {}
    except Exception:
        current_data = {}

    turning_on_multi = (
        updates.get("multi_location") is True
        and not current_data.get("multi_location")
    )

    supabase.table("companies").update(updates).eq("id", company_id).execute()

    if turning_on_multi:
        _seed_initial_location(company_id, current_data)

    try:
        res = supabase.table("companies").select("*").eq("id", company_id).single().execute()
    except Exception:
        raise HTTPException(404, "Company not found")
    if not res.data:
        raise HTTPException(404, "Company not found")

    return {"message": "Company settings updated", "data": res.data}


def _seed_initial_location(company_id: str, company_data: dict) -> None:
    """When a company first switches to multi_location, create one initial
    location from its existing lat/lng and auto-assign every current employee
    so that no one is locked out of clocking in."""
    # Skip if there's already a location for this company (idempotent)
    existing = (
        supabase.table("locations")
        .select("id")
        .eq("company_id", company_id)
        .limit(1)
        .execute()
    )
    if existing.data:
        return

    if company_data.get("lat") is None or company_data.get("lng") is None:
        # No existing single-point to migrate — admin will create locations manually
        return

    try:
        new_loc = supabase.table("locations").insert({
            "company_id":    company_id,
            "name":          "Kantor Utama",
            "lat":           company_data["lat"],
            "lng":           company_data["lng"],
            "radius_meters": company_data.get("radius_meters") or 100,
            "is_active":     True,
        }).execute()
        loc_id = (new_loc.data or [{}])[0].get("id")
    except Exception:
        return

    if not loc_id:
        return

    # Auto-assign all employees of this company to the new location
    emps = (
        supabase.table("profiles")
        .select("id")
        .eq("company_id", company_id)
        .execute()
    )
    rows = [{"user_id": e["id"], "location_id": loc_id} for e in (emps.data or [])]
    if rows:
        try:
            supabase.table("employee_locations").insert(rows).execute()
        except Exception:
            pass  # Some rows may already exist — non-fatal


@router.get("/report/monthly")
async def monthly_report(
    year: int = None,
    month: int = None,
    admin: dict = Depends(require_admin),
):
    """Per-employee monthly summary: hadir, cuti, lembur, total jam kerja."""
    from datetime import datetime
    from zoneinfo import ZoneInfo
    import calendar

    today = datetime.now(ZoneInfo("Asia/Jakarta")).date()
    year  = year  or today.year
    month = month or today.month

    first_day = f"{year}-{month:02d}-01"
    last_day  = f"{year}-{month:02d}-{calendar.monthrange(year, month)[1]:02d}"
    company_id = admin["company_id"]

    # All active employees
    emp_res = supabase.table("profiles").select("id, full_name, position").eq("company_id", company_id).eq("is_active", True).order("full_name").execute()
    employees = emp_res.data or []

    # All attendance in range
    att_res = supabase.table("attendance").select("id, user_id, status, clock_in, clock_out, break_start, break_end").eq("company_id", company_id).gte("date", first_day).lte("date", last_day).execute()
    att_rows = att_res.data or []

    # Approved corrections — overlay corrected clock_in/clock_out
    att_ids_view = [r["id"] for r in att_rows if r.get("id")]
    corr_map_view: dict = {}
    if att_ids_view:
        _c = supabase.table("attendance_corrections").select("attendance_id, requested_clock_in, requested_clock_out").eq("company_id", company_id).eq("status", "approved").in_("attendance_id", att_ids_view).execute()
        corr_map_view = {c["attendance_id"]: c for c in (_c.data or [])}
    for r in att_rows:
        c = corr_map_view.get(r.get("id"))
        if c:
            if c.get("requested_clock_in"):  r["clock_in"]  = c["requested_clock_in"]
            if c.get("requested_clock_out"): r["clock_out"] = c["requested_clock_out"]

    # Approved leaves in range
    leave_res = supabase.table("leave_requests").select("user_id, days_count").eq("company_id", company_id).eq("status", "approved").lte("start_date", last_day).gte("end_date", first_day).execute()
    leave_rows = leave_res.data or []

    # Approved overtime in range
    ot_res = supabase.table("overtime_requests").select("user_id, duration_minutes").eq("company_id", company_id).eq("status", "approved").gte("date", first_day).lte("date", last_day).execute()
    ot_rows = ot_res.data or []

    # Build per-user index
    from collections import defaultdict
    att_by_user   = defaultdict(list)
    leave_by_user = defaultdict(int)
    ot_by_user    = defaultdict(int)

    for r in att_rows:
        att_by_user[r["user_id"]].append(r)
    for r in leave_rows:
        leave_by_user[r["user_id"]] += r.get("days_count") or 0
    for r in ot_rows:
        ot_by_user[r["user_id"]] += r.get("duration_minutes") or 0

    result = []
    for emp in employees:
        uid  = emp["id"]
        rows = att_by_user[uid]

        hadir = sum(1 for r in rows if r.get("status") in ("present", "late", "early_leave"))
        work_min = sum(effective_work_minutes(r) for r in rows)

        result.append({
            "user_id":        uid,
            "full_name":      emp["full_name"],
            "position":       emp.get("position"),
            "hadir":          hadir,
            "cuti_days":      leave_by_user[uid],
            "lembur_minutes": ot_by_user[uid],
            "work_minutes":   work_min,
        })

    return {"year": year, "month": month, "data": result}


@router.get("/report/monthly/export")
async def export_monthly_report_xlsx(
    year: int = None,
    month: int = None,
    admin: dict = Depends(require_admin),
):
    """Export monthly per-employee summary as XLSX."""
    from datetime import datetime
    from zoneinfo import ZoneInfo
    import calendar
    from collections import defaultdict

    today = datetime.now(ZoneInfo("Asia/Jakarta")).date()
    year  = year  or today.year
    month = month or today.month

    first_day = f"{year}-{month:02d}-01"
    last_day  = f"{year}-{month:02d}-{calendar.monthrange(year, month)[1]:02d}"
    company_id = admin["company_id"]

    emp_res = supabase.table("profiles").select("id, full_name, position").eq("company_id", company_id).eq("is_active", True).order("full_name").execute()
    employees = emp_res.data or []

    att_res = supabase.table("attendance").select("id, user_id, status, clock_in, clock_out, break_start, break_end").eq("company_id", company_id).gte("date", first_day).lte("date", last_day).execute()
    att_rows = att_res.data or []

    att_ids_exp = [r["id"] for r in att_rows if r.get("id")]
    corr_map_exp: dict = {}
    if att_ids_exp:
        _ce = supabase.table("attendance_corrections").select("attendance_id, requested_clock_in, requested_clock_out").eq("company_id", company_id).eq("status", "approved").in_("attendance_id", att_ids_exp).execute()
        corr_map_exp = {c["attendance_id"]: c for c in (_ce.data or [])}
    for r in att_rows:
        c = corr_map_exp.get(r.get("id"))
        if c:
            if c.get("requested_clock_in"):  r["clock_in"]  = c["requested_clock_in"]
            if c.get("requested_clock_out"): r["clock_out"] = c["requested_clock_out"]

    leave_res = supabase.table("leave_requests").select("user_id, days_count").eq("company_id", company_id).eq("status", "approved").lte("start_date", last_day).gte("end_date", first_day).execute()
    leave_rows = leave_res.data or []

    ot_res = supabase.table("overtime_requests").select("user_id, duration_minutes").eq("company_id", company_id).eq("status", "approved").gte("date", first_day).lte("date", last_day).execute()
    ot_rows = ot_res.data or []

    att_by_user   = defaultdict(list)
    leave_by_user = defaultdict(int)
    ot_by_user    = defaultdict(int)
    for r in att_rows:
        att_by_user[r["user_id"]].append(r)
    for r in leave_rows:
        leave_by_user[r["user_id"]] += r.get("days_count") or 0
    for r in ot_rows:
        ot_by_user[r["user_id"]] += r.get("duration_minutes") or 0

    headers = ["Nama", "Jabatan", "Hadir", "Cuti (hari)", "Lembur (jam)", "Total Kerja (jam)"]
    rows = []
    for emp in employees:
        uid  = emp["id"]
        recs = att_by_user[uid]
        hadir    = sum(1 for r in recs if r.get("status") in ("present", "late", "early_leave"))
        work_min = sum(effective_work_minutes(r) for r in recs)
        rows.append([
            emp["full_name"],
            emp.get("position") or "",
            hadir,
            leave_by_user[uid],
            round(ot_by_user[uid] / 60, 1),
            round(work_min / 60, 1),
        ])

    buf = _make_xlsx(f"Rekap {year}-{month:02d}", headers, rows, "059669")
    return _xlsx_response(buf, f"rekap_{year}_{month:02d}.xlsx")


@router.get("/attendance/export")
async def export_attendance_xlsx(
    date_from: str = None,
    date_to: str = None,
    admin: dict = Depends(require_admin),
):
    """Export attendance data as XLSX. Defaults to current month."""
    from datetime import datetime as _dt
    from zoneinfo import ZoneInfo as _Zi
    import calendar

    if not date_from or not date_to:
        today = _dt.now(_Zi("Asia/Jakarta")).date()
        date_from = today.replace(day=1).isoformat()
        date_to = today.replace(day=calendar.monthrange(today.year, today.month)[1]).isoformat()

    res = (
        supabase.table("attendance")
        .select("id, date, clock_in, clock_out, break_start, break_end, clock_in_distance_m, status, notes, profiles(full_name, position), locations(name)")
        .eq("company_id", admin["company_id"])
        .gte("date", date_from)
        .lte("date", date_to)
        .order("date", desc=False)
        .order("clock_in", desc=False)
        .execute()
    )

    att_ids = [r["id"] for r in (res.data or []) if r.get("id")]
    corr_map: dict = {}
    if att_ids:
        corr_res = (
            supabase.table("attendance_corrections")
            .select("attendance_id, requested_clock_in, requested_clock_out")
            .eq("company_id", admin["company_id"])
            .eq("status", "approved")
            .in_("attendance_id", att_ids)
            .execute()
        )
        corr_map = {c["attendance_id"]: c for c in (corr_res.data or [])}

    def fmt(ts):
        if not ts:
            return ""
        try:
            dt = _dt.fromisoformat(ts.replace("Z", "+00:00"))
            return dt.astimezone(_Zi("Asia/Jakarta")).strftime("%H:%M")
        except Exception:
            return ts

    headers = ["Nama", "Jabatan", "Tanggal", "Jam Masuk", "Mulai Istirahat", "Selesai Istirahat", "Jam Keluar", "Durasi Kerja (jam)", "Lokasi", "Jarak (m)", "Status", "Catatan"]
    rows = []
    for row in (res.data or []):
        profile = row.get("profiles") or {}
        loc     = row.get("locations") or {}
        corr    = corr_map.get(row.get("id"))
        clock_in  = corr["requested_clock_in"]  if corr and corr.get("requested_clock_in")  else row.get("clock_in")
        clock_out = corr["requested_clock_out"] if corr and corr.get("requested_clock_out") else row.get("clock_out")
        work_min = effective_work_minutes({**row, "clock_in": clock_in, "clock_out": clock_out})
        rows.append([
            profile.get("full_name", ""),
            profile.get("position", ""),
            row.get("date", ""),
            fmt(clock_in),
            fmt(row.get("break_start")),
            fmt(row.get("break_end")),
            fmt(clock_out),
            round(work_min / 60, 1) if work_min else "",
            loc.get("name", ""),
            row.get("clock_in_distance_m", ""),
            _STATUS_ID.get(row.get("status", ""), row.get("status", "")),
            row.get("notes", ""),
        ])

    buf = _make_xlsx("Absensi", headers, rows, "2563EB")
    return _xlsx_response(buf, f"absensi_{date_from}_{date_to}.xlsx")


@router.patch("/attendance/{record_id}")
async def correct_attendance(
    record_id: str,
    body: AttendanceCorrectionRequest,
    admin: dict = Depends(require_admin),
):
    """Admin corrects clock-in / clock-out time for any attendance record in their company."""
    from datetime import datetime, timezone
    from zoneinfo import ZoneInfo

    try:
        rec = supabase.table("attendance").select("company_id, date, clock_in, clock_out, break_start, break_end").eq("id", record_id).single().execute()
    except Exception:
        raise HTTPException(404, "Attendance record not found")
    if not rec.data or rec.data["company_id"] != admin["company_id"]:
        raise HTTPException(404, "Attendance record not found")

    updates: dict = {}

    def parse_local(dt_str: str) -> str:
        """Accept 'YYYY-MM-DDTHH:MM' (naive, treated as WIB) and return UTC ISO string."""
        try:
            naive = datetime.fromisoformat(dt_str)
            wib = naive.replace(tzinfo=ZoneInfo("Asia/Jakarta"))
            return wib.astimezone(timezone.utc).isoformat()
        except Exception:
            raise HTTPException(400, f"Invalid datetime format: {dt_str!r}. Use YYYY-MM-DDTHH:MM")

    if body.clock_in is not None:
        updates["clock_in"] = parse_local(body.clock_in)
    if body.clock_out is not None:
        updates["clock_out"] = parse_local(body.clock_out)
    if body.break_start is not None:
        updates["break_start"] = parse_local(body.break_start) if body.break_start else None
    if body.break_end is not None:
        updates["break_end"] = parse_local(body.break_end) if body.break_end else None
    if body.notes is not None:
        updates["notes"] = body.notes

    if not updates:
        raise HTTPException(400, "Nothing to update")

    # Validate ordering: clock_out > clock_in, and break within that window.
    # Use `in updates` (not `or`) so that explicitly-cleared values (None) are not
    # overridden by the existing record's value during validation.
    final_in  = updates["clock_in"]    if "clock_in"    in updates else rec.data.get("clock_in")
    final_out = updates["clock_out"]   if "clock_out"   in updates else rec.data.get("clock_out")
    final_bs  = updates["break_start"] if "break_start" in updates else rec.data.get("break_start")
    final_be  = updates["break_end"]   if "break_end"   in updates else rec.data.get("break_end")

    if final_in and final_out and final_out <= final_in:
        raise HTTPException(400, "Jam keluar harus setelah jam masuk")
    if final_bs and final_be and final_be <= final_bs:
        raise HTTPException(400, "Jam selesai istirahat harus setelah jam mulai istirahat")
    if final_in and final_bs and final_bs < final_in:
        raise HTTPException(400, "Jam mulai istirahat harus setelah jam masuk")
    if final_out and final_be and final_be > final_out:
        raise HTTPException(400, "Jam selesai istirahat harus sebelum jam keluar")

    supabase.table("attendance").update(updates).eq("id", record_id).execute()
    return {"message": "Attendance corrected"}


@router.patch("/employees/{user_id}/role")
async def update_employee_role(
    user_id: str,
    role: str,
    admin: dict = Depends(require_admin),
):
    if role not in ("employee", "admin"):
        raise HTTPException(400, "role must be 'employee' or 'admin'")

    try:
        emp = supabase.table("profiles").select("company_id").eq("id", user_id).single().execute()
    except Exception:
        raise HTTPException(404, "Employee not found")
    if not emp.data or emp.data["company_id"] != admin["company_id"]:
        raise HTTPException(404, "Employee not found")
    if user_id == admin["id"]:
        raise HTTPException(400, "Tidak dapat mengubah role diri sendiri")

    supabase.table("profiles").update({"role": role}).eq("id", user_id).execute()
    return {"message": f"Role updated to {role}"}


@router.patch("/employees/{user_id}/active")
async def toggle_employee_active(
    user_id: str,
    is_active: bool,
    admin: dict = Depends(require_admin),
):
    try:
        emp = supabase.table("profiles").select("company_id,role").eq("id", user_id).single().execute()
    except Exception:
        raise HTTPException(404, "Employee not found")
    if not emp.data or emp.data["company_id"] != admin["company_id"]:
        raise HTTPException(404, "Employee not found")
    if user_id == admin["id"]:
        raise HTTPException(400, "Cannot deactivate yourself")

    supabase.table("profiles").update({"is_active": is_active}).eq("id", user_id).execute()
    return {"message": "activated" if is_active else "deactivated"}


@router.get("/leave/export")
async def export_leave_xlsx(
    status_filter: str = None,
    date_from: str = None,
    date_to: str = None,
    admin: dict = Depends(require_admin),
):
    """Export leave requests as XLSX."""
    if date_from and date_to and date_from > date_to:
        raise HTTPException(400, "date_from tidak boleh setelah date_to")
    q = (
        supabase.table("leave_requests")
        .select("*, profiles!leave_requests_user_id_fkey(full_name, position)")
        .eq("company_id", admin["company_id"])
        .order("start_date", desc=True)
    )
    if status_filter:
        q = q.eq("status", status_filter)
    if date_from:
        q = q.gte("start_date", date_from)
    if date_to:
        q = q.lte("start_date", date_to)
    res = q.execute()

    headers = ["Nama", "Jabatan", "Jenis Cuti", "Tanggal Mulai", "Tanggal Selesai", "Jumlah Hari", "Status", "Alasan", "Catatan Reviewer"]
    rows = []
    for row in (res.data or []):
        profile = row.get("profiles") or {}
        rows.append([
            profile.get("full_name", ""),
            profile.get("position", ""),
            _LEAVE_TYPE_ID.get(row.get("leave_type", ""), row.get("leave_type", "")),
            row.get("start_date", ""),
            row.get("end_date", ""),
            row.get("days_count", ""),
            _STATUS_ID.get(row.get("status", ""), row.get("status", "")),
            row.get("reason", ""),
            row.get("reviewer_note", ""),
        ])

    suffix = f"_{date_from}_{date_to}" if date_from and date_to else ""
    buf = _make_xlsx("Cuti", headers, rows, "DC2626")
    return _xlsx_response(buf, f"cuti{suffix}.xlsx")


@router.get("/overtime/export")
async def export_overtime_xlsx(
    status_filter: str = None,
    date_from: str = None,
    date_to: str = None,
    admin: dict = Depends(require_admin),
):
    """Export overtime requests as XLSX."""
    q = (
        supabase.table("overtime_requests")
        .select("*, profiles!overtime_requests_user_id_fkey(full_name, position)")
        .eq("company_id", admin["company_id"])
        .order("date", desc=True)
    )
    if status_filter:
        q = q.eq("status", status_filter)
    if date_from:
        q = q.gte("date", date_from)
    if date_to:
        q = q.lte("date", date_to)
    res = q.execute()

    headers = ["Nama", "Jabatan", "Tanggal", "Jam Mulai", "Jam Selesai", "Durasi (jam)", "Status", "Keterangan"]
    rows = []
    for row in (res.data or []):
        profile = row.get("profiles") or {}
        dur_min = row.get("duration_minutes") or 0
        rows.append([
            profile.get("full_name", ""),
            profile.get("position", ""),
            row.get("date", ""),
            (row.get("start_time") or "")[:5],
            (row.get("end_time") or "")[:5],
            round(dur_min / 60, 1) if dur_min else "",
            _STATUS_ID.get(row.get("status", ""), row.get("status", "")),
            row.get("description", ""),
        ])

    suffix = f"_{date_from}_{date_to}" if date_from and date_to else ""
    buf = _make_xlsx("Lembur", headers, rows, "D97706")
    return _xlsx_response(buf, f"lembur{suffix}.xlsx")


@router.get("/employees/{user_id}/attendance/export")
async def export_employee_attendance_xlsx(
    user_id: str,
    admin: dict = Depends(require_admin),
):
    """Export all attendance records for a single employee as XLSX."""
    from datetime import datetime as _dt
    from zoneinfo import ZoneInfo as _Zi

    try:
        emp = supabase.table("profiles").select("company_id, full_name").eq("id", user_id).single().execute()
    except Exception:
        raise HTTPException(404, "Karyawan tidak ditemukan")
    if not emp.data or emp.data["company_id"] != admin["company_id"]:
        raise HTTPException(404, "Karyawan tidak ditemukan")

    res = (
        supabase.table("attendance")
        .select("id, date, clock_in, clock_out, break_start, break_end, status, notes, locations(name)")
        .eq("user_id", user_id)
        .eq("company_id", admin["company_id"])
        .order("date", desc=False)
        .execute()
    )

    emp_att_ids = [r["id"] for r in (res.data or []) if r.get("id")]
    emp_corr_map: dict = {}
    if emp_att_ids:
        _ec = supabase.table("attendance_corrections").select("attendance_id, requested_clock_in, requested_clock_out").eq("company_id", admin["company_id"]).eq("status", "approved").in_("attendance_id", emp_att_ids).execute()
        emp_corr_map = {c["attendance_id"]: c for c in (_ec.data or [])}

    def fmt(ts):
        if not ts:
            return ""
        try:
            dt = _dt.fromisoformat(ts.replace("Z", "+00:00"))
            return dt.astimezone(_Zi("Asia/Jakarta")).strftime("%H:%M")
        except Exception:
            return ts

    emp_name = emp.data["full_name"]
    headers = ["Tanggal", "Jam Masuk", "Mulai Istirahat", "Selesai Istirahat", "Jam Keluar", "Durasi Kerja (jam)", "Lokasi", "Status", "Catatan"]
    rows = []
    for row in (res.data or []):
        loc  = row.get("locations") or {}
        corr = emp_corr_map.get(row.get("id"))
        clock_in  = corr["requested_clock_in"]  if corr and corr.get("requested_clock_in")  else row.get("clock_in")
        clock_out = corr["requested_clock_out"] if corr and corr.get("requested_clock_out") else row.get("clock_out")
        work_min = effective_work_minutes({**row, "clock_in": clock_in, "clock_out": clock_out})
        rows.append([
            row.get("date", ""),
            fmt(clock_in),
            fmt(row.get("break_start")),
            fmt(row.get("break_end")),
            fmt(clock_out),
            round(work_min / 60, 1) if work_min else "",
            loc.get("name", ""),
            _STATUS_ID.get(row.get("status", ""), row.get("status", "")),
            row.get("notes", ""),
        ])

    safe_name = emp_name.replace(" ", "_")
    buf = _make_xlsx(emp_name[:31], headers, rows, "7C3AED")
    return _xlsx_response(buf, f"absensi_{safe_name}.xlsx")
